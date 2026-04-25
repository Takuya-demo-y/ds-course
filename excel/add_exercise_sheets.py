"""
各演習Excelファイルの「演習説明」シートに解説テキストを追加するスクリプト
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ─── スタイル定義 ───────────────────────────────────────────

# 見出しカラー
COLOR_H1_BG   = "1F3864"  # 濃紺（大見出し背景）
COLOR_H2_BG   = "2E75B6"  # 青（中見出し背景）
COLOR_H3_BG   = "BDD7EE"  # 薄青（小見出し背景）
COLOR_CODE_BG = "F2F2F2"  # 薄グレー（コード背景）
COLOR_TIP_BG  = "E2EFDA"  # 薄緑（ポイント背景）
COLOR_WARN_BG = "FFF2CC"  # 薄黄（注意背景）
COLOR_WHITE   = "FFFFFF"
COLOR_BLACK   = "000000"
COLOR_H3_TEXT = "1F3864"  # 濃紺テキスト

def h1(ws, row, text):
    """大見出し（演習タイトルなど）"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="メイリオ", bold=True, size=14, color=COLOR_WHITE)
    cell.fill = PatternFill("solid", fgColor=COLOR_H1_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 28
    return row + 1

def h2(ws, row, text):
    """中見出し（セクション）"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="メイリオ", bold=True, size=11, color=COLOR_WHITE)
    cell.fill = PatternFill("solid", fgColor=COLOR_H2_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1

def h3(ws, row, text):
    """小見出し（ステップなど）"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="メイリオ", bold=True, size=10, color=COLOR_H3_TEXT)
    cell.fill = PatternFill("solid", fgColor=COLOR_H3_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1

def body(ws, row, text, indent=1):
    """通常テキスト"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="メイリオ", size=10, color=COLOR_BLACK)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=indent)
    ws.row_dimensions[row].height = 18
    return row + 1

def code(ws, row, text):
    """コード行（グレー背景・等幅フォント）"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Courier New", size=9, color="1F3864")
    cell.fill = PatternFill("solid", fgColor=COLOR_CODE_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=2)
    ws.row_dimensions[row].height = 17
    return row + 1

def tip(ws, row, text):
    """ポイント・コツ（緑背景）"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="メイリオ", size=10, bold=True, color="375623")
    cell.fill = PatternFill("solid", fgColor=COLOR_TIP_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1

def warn(ws, row, text):
    """注意（黄背景）"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="メイリオ", size=10, bold=True, color="7F6000")
    cell.fill = PatternFill("solid", fgColor=COLOR_WARN_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1

def blank(ws, row, height=8):
    """空行"""
    ws.row_dimensions[row].height = height
    return row + 1

def setup_sheet(ws):
    """シートの列幅などを設定"""
    ws.column_dimensions["A"].width = 90
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════
# 演習1：デバッグ演習
# ═══════════════════════════════════════════════════════════════

def write_debug_sheet(ws):
    setup_sheet(ws)
    r = 1

    r = h1(ws, r, "  演習1　デバッグ演習　──　バグを見つけて修正しよう")
    r = blank(ws, r)

    # 目的
    r = h2(ws, r, "  この演習の目的")
    r = body(ws, r, "「VBAコード_バグあり.txt」のコードには、意図的にバグ（間違い）が2箇所仕込まれています。")
    r = body(ws, r, "デバッグの3つの道具を使って、バグを見つけ・修正し、正しい結果を出してください。")
    r = blank(ws, r)
    r = body(ws, r, "【正解の確認方法】　D2:D11 を選択したとき、画面右下の合計が「¥2,670,000」になれば正解です。")
    r = blank(ws, r)

    # シートの確認
    r = h2(ws, r, "  シートの構成を確認する")
    r = body(ws, r, "「売上データ」シートを開いてください。以下の列構成になっています。")
    r = blank(ws, r)
    r = body(ws, r, "  A列：商品名　　B列：数量　　C列：単価　　D列：売上金額（ここに計算結果を書き込む）")
    r = blank(ws, r)
    r = body(ws, r, "1行目はヘッダー行（見出し）です。データは2行目から始まります。")
    r = blank(ws, r)

    # デバッグの道具
    r = h2(ws, r, "  デバッグの3つの道具　──　まずここを覚えよう")
    r = blank(ws, r)

    r = h3(ws, r, "  道具1：Debug.Print（デバッグ・プリント）")
    r = body(ws, r, "コードが動いている最中に、変数の値を「イミディエイトウィンドウ」に書き出す命令です。")
    r = body(ws, r, "処理を止めずに確認できるので、ループ（繰り返し処理）の中での確認に最適です。")
    r = blank(ws, r)
    r = body(ws, r, "書き方の例：")
    r = code(ws, r, '  Debug.Print "行:" & i & "  商品名:" & Cells(i, 1).Value')
    r = blank(ws, r)
    r = body(ws, r, "コード内の「' Debug.Print」の行頭の「'」を消すと有効になります。")
    r = body(ws, r, "確認が終わったら、また「'」を付けてコメントアウトに戻してください。")
    r = blank(ws, r)

    r = h3(ws, r, "  イミディエイトウィンドウ の開き方")
    r = body(ws, r, "Debug.Print の結果が表示される「小窓」です。")
    r = blank(ws, r)
    r = body(ws, r, "  ① VBE（Visual Basic Editor）を開く　→　キーボードで「Alt + F11」")
    r = body(ws, r, "  ② VBE が開いたら「Ctrl + G」を押す")
    r = body(ws, r, "  ③ 画面の下部に「イミディエイト」という小窓が表示される")
    r = blank(ws, r)
    r = tip(ws, r, "  ポイント：Debug.Print を実行したあと Ctrl+G で確認しましょう。")
    r = blank(ws, r)

    r = h3(ws, r, "  道具2：ブレークポイント（F9）")
    r = body(ws, r, "指定した行でコードの実行を一時停止させる機能です。停止中に変数の値をマウスで確認できます。")
    r = blank(ws, r)
    r = body(ws, r, "  ① 止めたい行にカーソルを置く")
    r = body(ws, r, "  ② F9 キーを押す（行の左端に赤丸 ● が表示される）")
    r = body(ws, r, "  ③ F5 でマクロを実行すると、その行でピタッと止まる")
    r = body(ws, r, "  ④ 止まっている状態で変数名にマウスをかざすと、今の値がポップアップで表示される")
    r = body(ws, r, "  ⑤ 解除するときはもう一度 F9 を押す")
    r = blank(ws, r)

    r = h3(ws, r, "  道具3：ステップ実行（F8）")
    r = body(ws, r, "F8 を1回押すたびに1行ずつコードを実行します。黄色のハイライトが現在実行中の行を示します。")
    r = blank(ws, r)
    r = body(ws, r, "  ① マクロ名にカーソルを置いた状態で F8 を押す")
    r = body(ws, r, "  ② Sub の1行目が黄色くなる（次に実行する行）")
    r = body(ws, r, "  ③ F8 を押すたびに1行ずつ進む")
    r = body(ws, r, "  ④ 変数にマウスをかざすと今の値が確認できる")
    r = blank(ws, r)
    r = tip(ws, r, "  合わせ技：F9でブレークポイントを設定して止めた後、F8で1行ずつ確認するのが最も効率的です。")
    r = blank(ws, r)

    # バグの説明
    r = h2(ws, r, "  バグの場所と修正方法")
    r = blank(ws, r)

    r = h3(ws, r, "  バグ1：ループの開始番号が間違っている")
    r = body(ws, r, "【バグのあるコード】")
    r = code(ws, r, "  For i = 1 To lastRow        ← 「1」になっている")
    r = blank(ws, r)
    r = body(ws, r, "【正しいコード】")
    r = code(ws, r, "  For i = 2 To lastRow        ← 「2」に変える")
    r = blank(ws, r)
    r = body(ws, r, "【なぜバグか】")
    r = body(ws, r, "1行目はヘッダー行（「商品名」「数量」…という見出し）です。i=1 から始めると、")
    r = body(ws, r, "「商品名」という文字を数量として計算しようとするため、結果がおかしくなります。")
    r = body(ws, r, "データは2行目から始まるので、For i = 2 から始めるのが正解です。")
    r = blank(ws, r)
    r = tip(ws, r, "  確認方法：Debug.Print のコメント「'」を外して実行。「行:1  商品名:商品名」と出れば1行目が処理されている証拠。")
    r = blank(ws, r)

    r = h3(ws, r, "  バグ2：書き込む列番号が間違っている")
    r = body(ws, r, "【バグのあるコード】")
    r = code(ws, r, "  Cells(i, 5).Value = 売上     ← 列番号が「5」（E列）になっている")
    r = blank(ws, r)
    r = body(ws, r, "【正しいコード】")
    r = code(ws, r, "  Cells(i, 4).Value = 売上     ← 列番号を「4」（D列）に変える")
    r = blank(ws, r)
    r = body(ws, r, "【列番号の対応表】　A列=1　B列=2　C列=3　D列=4　E列=5")
    r = body(ws, r, "売上金額はD列（4列目）に書き込む必要があります。5と書いてしまうとE列に入ります。")
    r = blank(ws, r)

    # 手順
    r = h2(ws, r, "  演習の進め方（手順）")
    r = body(ws, r, "  1. VBE を開く（Alt + F11）")
    r = body(ws, r, "  2. 「VBAコード_バグあり.txt」の内容をコピーして標準モジュールに貼り付ける")
    r = body(ws, r, "  3. F5 でマクロを実行して、何が起きるか確認する")
    r = body(ws, r, "  4. Debug.Print のコメント「'」を外して、イミディエイトウィンドウで変数を確認する")
    r = body(ws, r, "  5. バグ1（For の開始番号）を修正して再実行する")
    r = body(ws, r, "  6. バグ2（Cells の列番号）を修正して再実行する")
    r = body(ws, r, "  7. D2:D11 の合計が ¥2,670,000 になれば完成！")
    r = blank(ws, r)
    r = body(ws, r, "【完成した解答コードは「VBAコード_解答.txt」に記載されています。】")


# ═══════════════════════════════════════════════════════════════
# 演習2：Function演習
# ═══════════════════════════════════════════════════════════════

def write_function_sheet(ws):
    setup_sheet(ws)
    r = 1

    r = h1(ws, r, "  演習2　Function演習　──　売上評価マクロを作ろう")
    r = blank(ws, r)

    # 目的
    r = h2(ws, r, "  この演習の目的")
    r = body(ws, r, "「Function（関数）」の作り方と使い方を学びます。")
    r = body(ws, r, "Function を使うと「同じ処理を1か所にまとめて、どこからでも呼び出せる」コードが書けます。")
    r = blank(ws, r)
    r = body(ws, r, "【最終ゴール】　E列に各行の売上評価（A / B / C）が自動で入力される")
    r = blank(ws, r)

    # シートの確認
    r = h2(ws, r, "  シートの構成を確認する")
    r = body(ws, r, "「売上データ」シートを開いてください。以下の列構成になっています。")
    r = blank(ws, r)
    r = body(ws, r, "  A列：担当者名　　B列：支店名　　C列：商品カテゴリ　　D列：売上金額　　E列：評価（ここに入力する）")
    r = blank(ws, r)

    # Function とは
    r = h2(ws, r, "  Function とは何か？　Sub との違い")
    r = blank(ws, r)

    r = h3(ws, r, "  Sub（サブ）：仕事をするだけ、答えを返さない")
    r = body(ws, r, "セルにデータを書き込む、メッセージを出す、など「やりっぱなし」の処理。")
    r = code(ws, r, "  Sub 処理名()")
    r = code(ws, r, "      ... 処理 ...")
    r = code(ws, r, "  End Sub")
    r = blank(ws, r)

    r = h3(ws, r, "  Function（ファンクション）：何かを調べて、答えを返す")
    r = body(ws, r, "「この売上金額は A/B/C のどれ？」と聞いたら「A です」と答えを返してくれるもの。")
    r = code(ws, r, "  Function 関数名(引数 As 型) As 返す型")
    r = code(ws, r, "      関数名 = 戻り値     ← 「関数名 = 値」の形で答えを返す")
    r = code(ws, r, "  End Function")
    r = blank(ws, r)
    r = warn(ws, r, "  注意：「関数名 = 値」の書き方は Function 専用です。End Function の前に必ず書きましょう。")
    r = blank(ws, r)

    # STEP 1
    r = h2(ws, r, "  STEP 1　売上評価 Function を作る")
    r = blank(ws, r)
    r = body(ws, r, "【仕様】")
    r = body(ws, r, "  ・引数として「売上金額」（数値）を受け取る")
    r = body(ws, r, "  ・1,000,000円以上　→　\"A\" を返す")
    r = body(ws, r, "  ・  500,000円以上　→　\"B\" を返す")
    r = body(ws, r, "  ・  500,000円未満　→　\"C\" を返す")
    r = blank(ws, r)

    r = body(ws, r, "【完成形のコード】")
    r = code(ws, r, "  Function 売上評価(売上金額 As Long) As String")
    r = code(ws, r, "      If 売上金額 >= 1000000 Then")
    r = code(ws, r, '          売上評価 = "A"')
    r = code(ws, r, "      ElseIf 売上金額 >= 500000 Then")
    r = code(ws, r, '          売上評価 = "B"')
    r = code(ws, r, "      Else")
    r = code(ws, r, '          売上評価 = "C"')
    r = code(ws, r, "      End If")
    r = code(ws, r, "  End Function")
    r = blank(ws, r)
    r = tip(ws, r, "  ポイント：ElseIf の条件は「大きい順（100万 → 50万）」に書く。逆に書くと全員Bになるバグになります。")
    r = blank(ws, r)
    r = body(ws, r, "【STEP 1 の単体テスト方法】")
    r = body(ws, r, "イミディエイトウィンドウ（Ctrl+G）を開いて、以下を入力して Enter を押す：")
    r = code(ws, r, "  ? 売上評価(1200000)")
    r = body(ws, r, "「A」と表示されれば STEP 1 は完成です。")
    r = blank(ws, r)

    # STEP 2
    r = h2(ws, r, "  STEP 2　最終行取得 Function を作る")
    r = blank(ws, r)
    r = body(ws, r, "【何をする Function か】")
    r = body(ws, r, "指定したシートの指定した列で「データが入っている最後の行番号」を返します。")
    r = body(ws, r, "データの件数が毎回変わっても、この Function を使えば自動で対応できます。")
    r = blank(ws, r)

    r = body(ws, r, "【完成形のコード】")
    r = code(ws, r, "  Function 最終行取得(ws As Worksheet, col As Long) As Long")
    r = code(ws, r, "      最終行取得 = ws.Cells(ws.Rows.Count, col).End(xlUp).Row")
    r = code(ws, r, "  End Function")
    r = blank(ws, r)
    r = body(ws, r, "【コードの意味】　ws.Cells(ws.Rows.Count, col).End(xlUp).Row")
    r = body(ws, r, "  = 「シートの一番下のセルから、上方向に Ctrl+↑ を押したときに止まる行番号」を取得する")
    r = blank(ws, r)
    r = body(ws, r, "【STEP 2 の単体テスト方法】")
    r = body(ws, r, "「売上データ」シートをアクティブにしてから、イミディエイトウィンドウで：")
    r = code(ws, r, "  ? 最終行取得(ActiveSheet, 4)")
    r = body(ws, r, "D列のデータ行数+1 の数字が表示されれば完成です（例：データが10件なら「11」）。")
    r = blank(ws, r)

    # STEP 3
    r = h2(ws, r, "  STEP 3　2つの Function を使う Sub を完成させる")
    r = blank(ws, r)
    r = body(ws, r, "【完成形のコード】")
    r = code(ws, r, "  Sub 評価を入力する()")
    r = code(ws, r, "      Dim ws      As Worksheet")
    r = code(ws, r, "      Dim i       As Long")
    r = code(ws, r, "      Dim lastRow As Long")
    r = code(ws, r, "      Dim sales   As Long")
    r = code(ws, r, "      Dim grade   As String")
    r = blank(ws, r)
    r = code(ws, r, '      Set ws = ThisWorkbook.Worksheets("売上データ")')
    r = code(ws, r, "      lastRow = 最終行取得(ws, 4)      ' D列を基準に最終行を取得")
    r = blank(ws, r)
    r = code(ws, r, "      For i = 2 To lastRow")
    r = code(ws, r, "          sales = ws.Cells(i, 4).Value    ' D列の売上金額を取得")
    r = code(ws, r, "          grade = 売上評価(sales)          ' 売上評価 Function で判定")
    r = code(ws, r, "          ws.Cells(i, 5).Value = grade    ' E列に評価を書き込む")
    r = code(ws, r, "      Next i")
    r = blank(ws, r)
    r = code(ws, r, '      MsgBox "評価の入力が完了しました！"')
    r = code(ws, r, "  End Sub")
    r = blank(ws, r)
    r = tip(ws, r, "  ポイント：Worksheet 型の変数に代入するときは「Set」が必要。忘れるとエラーになります。")
    r = blank(ws, r)

    # 進め方
    r = h2(ws, r, "  演習の進め方（手順）")
    r = body(ws, r, "  1. VBE を開く（Alt + F11）")
    r = body(ws, r, "  2. 「VBAコード_テンプレート.txt」の内容をコピーして標準モジュールに貼り付ける")
    r = body(ws, r, "  3. STEP 1　売上評価 Function の TODO を埋める")
    r = body(ws, r, "  4. イミディエイトウィンドウで単体テストする（? 売上評価(1200000) → \"A\"）")
    r = body(ws, r, "  5. STEP 2　最終行取得 Function の TODO を埋める")
    r = body(ws, r, "  6. STEP 3　Sub の TODO を2か所埋める（lastRow = ？　と　grade = ？）")
    r = body(ws, r, "  7. 「評価を入力する」Sub を実行して E列に A/B/C が入れば完成！")
    r = blank(ws, r)
    r = body(ws, r, "【完成した解答コードは「VBAコード_解答.txt」に記載されています。】")


# ═══════════════════════════════════════════════════════════════
# 演習3：シート操作演習
# ═══════════════════════════════════════════════════════════════

def write_sheet_ops_sheet(ws):
    setup_sheet(ws)
    r = 1

    r = h1(ws, r, "  演習3　シート操作演習　──　月別シート自動振り分けマクロを作ろう")
    r = blank(ws, r)

    # 目的
    r = h2(ws, r, "  この演習の目的")
    r = body(ws, r, "「全データ」シートに入っている1年分の売上データを、月ごとに別シートへ自動で振り分けます。")
    r = body(ws, r, "シートを動的に作成する方法と、日付から月番号を取り出す方法を学びます。")
    r = blank(ws, r)
    r = body(ws, r, "【最終ゴール】　「1月」〜「12月」のシートが自動生成され、各月のデータが正しく振り分けられる")
    r = blank(ws, r)

    # シートの確認
    r = h2(ws, r, "  シートの構成を確認する")
    r = body(ws, r, "「全データ」シートを開いてください。以下の列構成になっています。")
    r = blank(ws, r)
    r = body(ws, r, "  A列：日付（例: 2024/01/15）　B列：支店名　C列：商品カテゴリ　D列：担当者　E列：売上金額")
    r = blank(ws, r)

    # 新しい概念
    r = h2(ws, r, "  このSTEPで登場する新しい概念")
    r = blank(ws, r)

    r = h3(ws, r, "  Set（セット）：シートなどの「もの」を変数に入れるときに必要")
    r = body(ws, r, "普通の変数（数値や文字列）への代入は Set は不要です。")
    r = body(ws, r, "シートやブックなどの「オブジェクト」を変数に入れるときだけ Set が必要です。")
    r = blank(ws, r)
    r = code(ws, r, "  Dim x As Long")
    r = code(ws, r, "  x = 10                                 ← 普通の変数は Set 不要")
    r = blank(ws, r)
    r = code(ws, r, "  Dim ws As Worksheet")
    r = code(ws, r, '  Set ws = Worksheets("集計")            ← オブジェクトは Set が必要！')
    r = blank(ws, r)
    r = warn(ws, r, "  注意：Set を忘れると「オブジェクト変数またはWith ブロック変数が設定されていません」エラーが出ます。")
    r = blank(ws, r)

    r = h3(ws, r, "  On Error Resume Next：シートの存在確認に使う特別な命令")
    r = body(ws, r, "「次の1行でエラーが起きても、止まらずに次へ進んでください」という命令です。")
    r = body(ws, r, "存在しないシート名を指定するとエラーになるため、この命令で一時的にエラーを無視します。")
    r = blank(ws, r)
    r = code(ws, r, "  On Error Resume Next          ← エラー無視を開始")
    r = code(ws, r, '  Set ws = Worksheets("1月")    ← シートがなければエラー → でも続行（ws = Nothing のまま）')
    r = code(ws, r, "  On Error GoTo 0               ← 必ずここで解除！")
    r = blank(ws, r)
    r = warn(ws, r, "  重要：On Error GoTo 0 を必ず書くこと。解除しないと、その後のエラーも全部無視されます。")
    r = blank(ws, r)

    r = h3(ws, r, "  Nothing（ナッシング）：変数が「空っぽ」の状態")
    r = body(ws, r, "オブジェクト変数に何も入っていない状態を「Nothing」と言います。")
    r = body(ws, r, "シートが存在しない場合、Set しても ws は Nothing のままになります。")
    r = blank(ws, r)
    r = code(ws, r, "  If Not ws Is Nothing Then     ← ws が空っぽでない（シートが存在する）")
    r = code(ws, r, "      ws.Delete")
    r = code(ws, r, "      Set ws = Nothing          ← 削除後、変数をリセット")
    r = code(ws, r, "  End If")
    r = blank(ws, r)

    r = h3(ws, r, "  Month() 関数：日付から月の数字を取り出す")
    r = body(ws, r, "日付型の値から「月番号だけ」を取り出す関数です。")
    r = blank(ws, r)
    r = code(ws, r, "  Month(CDate(\"2024/03/15\"))    → 3  （3月）")
    r = code(ws, r, "  Month(CDate(\"2024/11/01\"))    → 11 （11月）")
    r = blank(ws, r)
    r = code(ws, r, "  monthNum  = Month(dateVal)         ' 月番号を取り出す")
    r = code(ws, r, '  sheetName = monthNum & "月"         \'  "3月" などのシート名を作る')
    r = blank(ws, r)

    # STEP 1
    r = h2(ws, r, "  STEP 1　月別シートを準備する（1月〜12月を自動生成）")
    r = blank(ws, r)
    r = body(ws, r, "【処理の流れ】")
    r = body(ws, r, "  ① 1 から 12 まで繰り返す（For i = 1 To 12）")
    r = body(ws, r, "  ② シート名を作る（\"1月\"〜\"12月\"）")
    r = body(ws, r, "  ③ 既に同名シートがあれば削除する（On Error Resume Next パターン）")
    r = body(ws, r, "  ④ 新しいシートを末尾に追加して名前をつける")
    r = body(ws, r, "  ⑤ ヘッダー行を書き込む")
    r = blank(ws, r)
    r = body(ws, r, "【完成形のコード（主要部分）】")
    r = code(ws, r, "  On Error Resume Next")
    r = code(ws, r, "  Set ws = Worksheets(sheetName)")
    r = code(ws, r, "  On Error GoTo 0")
    r = blank(ws, r)
    r = code(ws, r, "  If Not ws Is Nothing Then")
    r = code(ws, r, "      ws.Delete")
    r = code(ws, r, "      Set ws = Nothing")
    r = code(ws, r, "  End If")
    r = blank(ws, r)
    r = code(ws, r, "  Worksheets.Add(After:=Worksheets(Worksheets.Count)).Name = sheetName")
    r = blank(ws, r)
    r = tip(ws, r, "  ポイント：シートを削除するときの確認ダイアログを止めるために Application.DisplayAlerts = False を使います。処理後は True に戻す！")
    r = blank(ws, r)

    # STEP 2
    r = h2(ws, r, "  STEP 2　データを月別シートに振り分ける")
    r = blank(ws, r)
    r = body(ws, r, "【処理の流れ】")
    r = body(ws, r, "  ① 「全データ」シートの2行目から最終行まで1行ずつ読む")
    r = body(ws, r, "  ② A列から日付を取得する")
    r = body(ws, r, "  ③ Month() で月番号を取り出す（例：3月なら 3）")
    r = body(ws, r, "  ④ 月番号 + \"月\" でシート名を作る（例：\"3月\"）")
    r = body(ws, r, "  ⑤ そのシートの次の空行を求めて、A〜E列のデータを転記する")
    r = blank(ws, r)
    r = body(ws, r, "【完成形のコード（主要部分）】")
    r = code(ws, r, "  dateVal   = srcWs.Cells(i, 1).Value    ' A列から日付を取得")
    r = code(ws, r, "  monthNum  = Month(dateVal)              ' 月番号を取り出す")
    r = code(ws, r, '  sheetName = monthNum & "月"             \' "3月" などのシート名を作る')
    r = blank(ws, r)
    r = code(ws, r, "  Set dstWs = Worksheets(sheetName)       ' 転記先シートを取得")
    r = code(ws, r, "  dstLastRow = dstWs.Cells(dstWs.Rows.Count, 1).End(xlUp).Row + 1  ' 次の空行")
    r = blank(ws, r)
    r = code(ws, r, "  For col = 1 To 5")
    r = code(ws, r, "      dstWs.Cells(dstLastRow, col).Value = srcWs.Cells(i, col).Value")
    r = code(ws, r, "  Next col")
    r = blank(ws, r)
    r = tip(ws, r, "  ポイント：大量データを転記するときは Application.ScreenUpdating = False で画面更新を止めると高速になります。処理後は True に戻す！")
    r = blank(ws, r)

    # 進め方
    r = h2(ws, r, "  演習の進め方（手順）")
    r = body(ws, r, "  1. VBE を開く（Alt + F11）")
    r = body(ws, r, "  2. 「VBAコード_テンプレート.txt」の内容をコピーして標準モジュールに貼り付ける")
    r = body(ws, r, "  3. STEP 1 の TODO を2箇所埋める（シート削除 / シート追加）")
    r = body(ws, r, "  4. 「月別シートを準備する」Sub を実行して 1月〜12月 のシートができているか確認する")
    r = body(ws, r, "  5. STEP 2 の TODO を埋める（転記先シートの取得・次の空行の計算・セルのコピー）")
    r = body(ws, r, "  6. 「月別振り分け実行」Sub を実行して各月シートにデータが入れば完成！")
    r = blank(ws, r)
    r = body(ws, r, "【完成した解答コードは「VBAコード_解答.txt」に記載されています。】")


# ═══════════════════════════════════════════════════════════════
# 演習4：総合演習
# ═══════════════════════════════════════════════════════════════

def write_comprehensive_sheet(ws):
    setup_sheet(ws)
    r = 1

    r = h1(ws, r, "  総合演習　支店別売上集計ツール　──　2日間の知識を組み合わせて完成させよう")
    r = blank(ws, r)

    # 目的
    r = h2(ws, r, "  この演習の目的")
    r = body(ws, r, "2日間で学んだすべての知識を組み合わせて、実務で使えるツールを完成させます。")
    r = body(ws, r, "「支店データ」フォルダにある 24 ファイル（4支店 × 6か月）を自動で集計します。")
    r = blank(ws, r)
    r = body(ws, r, "【最終ゴール】　「集計結果」シートに24件分の支店・年月・売上合計が一覧表示される")
    r = blank(ws, r)

    # 処理の全体像
    r = h2(ws, r, "  処理の全体像")
    r = blank(ws, r)
    r = body(ws, r, "  ① ボタンを押す（または マクロを実行）")
    r = body(ws, r, "  ② 「支店データ」フォルダの .xlsx ファイルを Dir() で一覧取得　←　Day 1 の知識")
    r = body(ws, r, "  ③ ファイルが実際に存在するか If Dir() で確認　←　Block 3 の知識")
    r = body(ws, r, "  ④ ファイルを開いて「最終行取得」Function で行数を取得　←　Block 2 の知識")
    r = body(ws, r, "  ⑤ 最終行の F列（月間合計）を取得して「集計結果」シートに転記")
    r = body(ws, r, "  ⑥ 24 件処理したら完了メッセージを表示")
    r = blank(ws, r)

    # TODO の説明
    r = h2(ws, r, "  TODO 一覧と解説")
    r = blank(ws, r)

    r = h3(ws, r, "  TODO-1（必須）：ファイルを開く前に存在確認を追加する")
    r = body(ws, r, "【なぜ必要か】")
    r = body(ws, r, "Dir() でファイル一覧を取得した後、そのファイルが移動・削除されている可能性があります。")
    r = body(ws, r, "確認せずに Workbooks.Open すると「ファイルが見つかりません」エラーになります。")
    r = blank(ws, r)
    r = body(ws, r, "【完成形のコード】")
    r = code(ws, r, "  If Dir(filePath) <> \"\" Then    ← ファイルが存在する場合だけ処理する")
    r = code(ws, r, "")
    r = code(ws, r, "      Set wb    = Workbooks.Open(filePath, ReadOnly:=True)")
    r = code(ws, r, "      Set srcWs = wb.Worksheets(\"売上明細\")")
    r = code(ws, r, "      ... （処理）...")
    r = code(ws, r, "      wb.Close SaveChanges:=False")
    r = code(ws, r, "")
    r = code(ws, r, "  End If                           ← If ブロックの終わり")
    r = blank(ws, r)
    r = tip(ws, r, "  ポイント：ReadOnly:=True で開くと、誤って上書きするリスクがなくなります。閉じるときも「保存しますか？」が出ません。")
    r = blank(ws, r)

    r = h3(ws, r, "  TODO-2（必須）：最終行取得 Function を呼び出す")
    r = body(ws, r, "【なぜ必要か】")
    r = body(ws, r, "各支店ファイルの明細行数は毎月変わります。最終行を自動で取得することで、")
    r = body(ws, r, "件数が変わってもコードを修正する必要がなくなります。")
    r = blank(ws, r)
    r = body(ws, r, "【完成形のコード】")
    r = code(ws, r, "  lastRow = 最終行取得(srcWs, 1)   ← A列を基準に最終行番号を取得")
    r = blank(ws, r)
    r = body(ws, r, "この1行だけです。Block 2 で作った Function がそのまま使えます。これが再利用の強みです。")
    r = blank(ws, r)

    r = h3(ws, r, "  TODO-3（推奨）：ScreenUpdating 等の最適化設定を有効にする")
    r = body(ws, r, "【なぜ必要か】")
    r = body(ws, r, "24ファイルを開いて閉じる処理を繰り返すと、画面がちらついて処理が遅くなります。")
    r = body(ws, r, "以下の設定を入れるだけで大幅に速くなります。")
    r = blank(ws, r)
    r = body(ws, r, "【完成形のコード】（コメントの「'」を外すだけ）")
    r = code(ws, r, "  Application.ScreenUpdating = False        ← 画面更新を止める")
    r = code(ws, r, "  Application.Calculation = xlCalculationManual  ← 自動計算を止める")
    r = code(ws, r, "  Application.DisplayAlerts = False         ← 確認ダイアログを非表示")
    r = blank(ws, r)
    r = code(ws, r, "  ... （処理）...")
    r = blank(ws, r)
    r = code(ws, r, "  Application.ScreenUpdating = True         ← 必ず元に戻す！")
    r = code(ws, r, "  Application.Calculation = xlCalculationAutomatic")
    r = code(ws, r, "  Application.DisplayAlerts = True")
    r = blank(ws, r)
    r = warn(ws, r, "  重要：ScreenUpdating = True の戻し忘れに注意。戻さないと画面が固まったままになります。")
    r = blank(ws, r)

    # Dir の2回使いの解説
    r = h2(ws, r, "  Dir() を2か所で使う理由（混乱しやすいポイント）")
    r = blank(ws, r)
    r = body(ws, r, "このコードでは Dir() が2か所に出てきます。それぞれ役割が違います。")
    r = blank(ws, r)
    r = body(ws, r, '  1か所目：fileName = Dir(folderPath & "*.xlsx")')
    r = body(ws, r, "　　フォルダの中の .xlsx ファイルを検索して、最初のファイル名を取得する。")
    r = blank(ws, r)
    r = body(ws, r, '  2か所目：If Dir(filePath) <> "" Then')
    r = body(ws, r, "　　取得したファイルのフルパスが実際に存在するかを確認する。")
    r = blank(ws, r)
    r = tip(ws, r, "  ポイント：ループの中で fileName = Dir() と書くと「次のファイル名」が取得できます（引数なし）。")
    r = blank(ws, r)

    # 進め方
    r = h2(ws, r, "  演習の進め方（手順）")
    r = body(ws, r, "  1. VBE を開く（Alt + F11）")
    r = body(ws, r, "  2. 「VBAコード_テンプレート.txt」の内容をコピーして標準モジュールに貼り付ける")
    r = body(ws, r, "  3. TODO-1 を完成させる（If Dir(filePath) <> \"\" Then ... End If を追加）")
    r = body(ws, r, "  4. F5 で実行して動作確認する")
    r = body(ws, r, "  5. TODO-2 を完成させる（lastRow = 最終行取得(srcWs, 1)）")
    r = body(ws, r, "  6. 再実行して集計結果シートを確認する")
    r = body(ws, r, "  7. TODO-3 のコメント「'」を外して最適化を有効にする")
    r = body(ws, r, "  8. 完了メッセージに「24 件のファイルを処理しました。」と出れば完成！")
    r = blank(ws, r)
    r = body(ws, r, "【動作確認のポイント】")
    r = body(ws, r, "  ・「集計結果」シートの A列に支店名が入っているか")
    r = body(ws, r, "  ・B列に「2024年01月」のような形式で年月が入っているか")
    r = body(ws, r, "  ・C列に月間合計金額（数値）が入っているか")
    r = body(ws, r, "  ・処理後も Excel が普通に操作できるか（画面が固まっていないか）")
    r = blank(ws, r)
    r = body(ws, r, "【完成した解答コードは「VBAコード_解答.txt」に記載されています。】")


# ═══════════════════════════════════════════════════════════════
# メイン処理
# ═══════════════════════════════════════════════════════════════

base = "/Users/takuyayoshihara/Documents/development/excel/演習用ファイル"
SHEET_NAME = "演習説明"

files = [
    (f"{base}/01_デバッグ演習/デバッグ演習.xlsx",      write_debug_sheet),
    (f"{base}/02_Function演習/売上評価データ.xlsx",     write_function_sheet),
    (f"{base}/03_シート操作演習/月次売上データ.xlsx",   write_sheet_ops_sheet),
    (f"{base}/04_総合演習/集計ツール_ベース.xlsx",      write_comprehensive_sheet),
]

for path, writer in files:
    print(f"処理中: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path)

    # 既存の「演習説明」シートを削除して再作成
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    # 一番左に挿入
    ws = wb.create_sheet(SHEET_NAME, 0)
    writer(ws)

    wb.save(path)
    print(f"  → 保存完了: {path}")

print("\n全ファイルの処理が完了しました。")
