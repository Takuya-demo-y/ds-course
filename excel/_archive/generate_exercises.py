#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
2日目研修 演習用Excelファイル生成スクリプト
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import date, timedelta
import random

random.seed(42)

BASE_DIR = "/Users/takuyayoshihara/Documents/development/excel/演習用ファイル"

# ===== スタイル定数 =====
HDR_COLOR   = "2E4057"
EVEN_COLOR  = "F5F7FA"
TOTAL_COLOR = "FFF3CD"
OK_COLOR    = "D4EDDA"
WARN_COLOR  = "FFF3CD"
ERR_COLOR   = "F8D7DA"


def hdr_font():   return Font(color="FFFFFF", bold=True, name="Yu Gothic UI", size=10)
def body_font():  return Font(name="Yu Gothic UI", size=10)
def title_font(): return Font(bold=True, name="Yu Gothic UI", size=14, color="2E4057")
def code_font():  return Font(name="Courier New", size=10, color="333333")
def note_font():  return Font(name="Yu Gothic UI", size=10, color="777777")


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header(cell, value):
    cell.value = value
    cell.font = hdr_font()
    cell.fill = PatternFill(start_color=HDR_COLOR, end_color=HDR_COLOR, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def apply_body(cell, value=None, align="center"):
    if value is not None:
        cell.value = value
    cell.font = body_font()
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center")


def stripe_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color=EVEN_COLOR, end_color=EVEN_COLOR, fill_type="solid"
        )


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def mkdir(path):
    os.makedirs(path, exist_ok=True)


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False


# ===== 演習1: デバッグ演習 =====
def create_debug_exercise():
    dir_path = os.path.join(BASE_DIR, "01_デバッグ演習")
    mkdir(dir_path)

    wb = openpyxl.Workbook()

    # ── 演習説明シート ──
    ws_intro = wb.active
    ws_intro.title = "演習説明"
    ws_intro.column_dimensions["A"].width = 3
    ws_intro.column_dimensions["B"].width = 85
    hide_gridlines(ws_intro)

    def put(row, col, val, fnt=None, fill_color=None):
        c = ws_intro.cell(row=row, column=col, value=val)
        c.font = fnt or body_font()
        if fill_color:
            c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        c.alignment = Alignment(vertical="center")
        return c

    put(2, 2, "【演習1】デバッグ入門", title_font())
    set_row_height(ws_intro, 2, 36)

    put(4, 2, "■ 演習の目的", Font(bold=True, name="Yu Gothic UI", size=11))
    put(5, 2, "VBAコードに仕込まれた2つのバグを、Debug.Print・ブレークポイント・ステップ実行を使って発見・修正してください。")

    put(7, 2, "■ 手順", Font(bold=True, name="Yu Gothic UI", size=11))
    steps = [
        "1. Alt + F11 でVBEを開き、Module1の「売上計算」プロシージャを確認する",
        "2. F8キーでステップ実行し、Debug.Printの出力をイミディエイトウィンドウで確認する",
        "3. 「売上データ」シートのD列に正しい売上金額が入らない原因を特定する",
        "4. バグを2箇所修正して、D列に正しい金額が表示されることを確認する",
        "5. 【確認】D2:D11 の合計が ¥2,670,000 になっていれば正解！",
    ]
    for i, s in enumerate(steps):
        put(8 + i, 2, s)

    put(14, 2, "■ ヒント", Font(bold=True, name="Yu Gothic UI", size=11))
    hints = [
        "・ バグは「ループの開始行」と「書き込む列番号」の2箇所にあります",
        "・ Debug.Print i, Cells(i, 1).Value をループ内に追加すると、何が起きるか確認できます",
        "・ F9でブレークポイントを設定し、変数の値を確認しながらステップ実行してみましょう",
    ]
    for i, h in enumerate(hints):
        c = put(15 + i, 2, h)
        c.font = note_font()

    # ── 売上データシート ──
    ws_data = wb.create_sheet("売上データ")
    headers = ["商品名", "数量（個）", "単価（円）", "売上金額（円）"]
    products = [
        ("商品A", 10, 15000), ("商品B", 25, 8000),  ("商品C", 5,  45000),
        ("商品D", 30, 3500),  ("商品E", 12, 22000), ("商品F", 8,  38000),
        ("商品G", 50, 1200),  ("商品H", 15, 12000), ("商品I", 20, 9800),
        ("商品J", 6,  55000),
    ]

    for col, h in enumerate(headers, 1):
        apply_header(ws_data.cell(row=1, column=col), h)
    set_row_height(ws_data, 1, 25)

    for row, (name, qty, price) in enumerate(products, 2):
        apply_body(ws_data.cell(row=row, column=1), name, "left")
        apply_body(ws_data.cell(row=row, column=2), qty)
        apply_body(ws_data.cell(row=row, column=3), price)
        ws_data.cell(row=row, column=3).number_format = "#,##0"
        c = ws_data.cell(row=row, column=4)
        c.value = "（マクロで計算）"
        c.font = Font(name="Yu Gothic UI", size=10, color="BBBBBB", italic=True)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if row % 2 == 0:
            stripe_row(ws_data, row, 1, 4)
        set_row_height(ws_data, row, 22)

    for col, w in zip(range(1, 5), [15, 12, 14, 18]):
        set_col_width(ws_data, col, w)
    ws_data.freeze_panes = "A2"

    wb.save(os.path.join(dir_path, "デバッグ演習.xlsx"))
    print("✓ 01_デバッグ演習/デバッグ演習.xlsx")

    # ── VBAコード（バグあり）──
    vba = """\
' ============================================================
' 【演習1】デバッグ演習 - 売上計算マクロ（バグが2箇所あります）
' ============================================================
' このコードには意図的にバグが2箇所仕込まれています。
' Debug.Print・ブレークポイント・ステップ実行を使って
' バグを見つけて修正してください。
'
' 【期待する動作】
' 「売上データ」シートの D列（売上金額）に
' B列（数量）× C列（単価）の結果を書き込む
'
' 【正解確認】
' D2:D11 の合計が ¥2,670,000 になっていれば正解です
' ============================================================

Sub 売上計算()

    Dim i       As Long
    Dim lastRow As Long
    Dim 数量    As Long
    Dim 単価    As Long
    Dim 売上    As Long

    Worksheets("売上データ").Activate

    lastRow = Cells(Rows.Count, 1).End(xlUp).Row

    For i = 1 To lastRow    ' ★ バグ1: ここに問題があります

        ' Debug.Print で変数の中身を確認してみましょう（コメントを外して使う）
        ' Debug.Print "行:" & i & "  商品名:" & Cells(i, 1).Value

        数量 = Cells(i, 2).Value
        単価 = Cells(i, 3).Value
        売上 = 数量 * 単価

        Cells(i, 5).Value = 売上  ' ★ バグ2: ここに問題があります

    Next i

    MsgBox "計算が完了しました！" & vbCrLf & _
           "D2:D11 の合計を確認してみましょう。"

End Sub

' ============================================================
' ヒント（どうしてもわからない場合のみ見てください）
' ============================================================
'
' ヒント1: ヘッダー行（1行目）はデータではないのに、
'          ループで処理されてしまっています。
'          For文の開始番号を確認してください。
'
' ヒント2: Cells(行, 列) の「列」の数字が
'          D列の列番号と合っているか確認してください。
'          （A=1, B=2, C=3, D=4, E=5）
'
' ============================================================
"""
    with open(os.path.join(dir_path, "VBAコード_バグあり.txt"), "w", encoding="utf-8") as f:
        f.write(vba)
    print("✓ 01_デバッグ演習/VBAコード_バグあり.txt")


# ===== 演習2: エラーハンドリング演習 =====
def create_error_handling_exercise():
    dir_path = os.path.join(BASE_DIR, "02_エラーハンドリング演習")
    branch_dir = os.path.join(dir_path, "支店データ")
    mkdir(dir_path)
    mkdir(branch_dir)

    staff_map = {
        "東京支店":   ["田中 太郎", "鈴木 花子", "佐藤 一郎"],
        "大阪支店":   ["山田 次郎", "中村 美咲", "小林 健太"],
        "名古屋支店": ["加藤 三郎", "伊藤 由美", "渡辺 誠"],
    }

    for branch, staff in staff_map.items():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "売上データ"

        headers = ["日付", "商品名", "担当者", "数量（個）", "売上金額（円）"]
        for col, h in enumerate(headers, 1):
            apply_header(ws.cell(row=1, column=col), h)
        set_row_height(ws, 1, 25)

        products = ["商品A", "商品B", "商品C", "商品D"]
        start = date(2024, 1, 7)
        row = 2
        for week in range(8):
            d = start + timedelta(weeks=week)
            for _ in range(3):
                product = random.choice(products)
                person  = random.choice(staff)
                qty     = random.randint(5, 50)
                price   = random.choice([3500, 8000, 15000, 22000])
                amount  = qty * price

                apply_body(ws.cell(row=row, column=1), d.strftime("%Y/%m/%d"))
                apply_body(ws.cell(row=row, column=2), product, "left")
                apply_body(ws.cell(row=row, column=3), person,  "left")
                apply_body(ws.cell(row=row, column=4), qty)
                apply_body(ws.cell(row=row, column=5), amount)
                ws.cell(row=row, column=5).number_format = "#,##0"
                if row % 2 == 0:
                    stripe_row(ws, row, 1, 5)
                set_row_height(ws, row, 22)
                row += 1

        # 合計行
        ws.cell(row=row, column=4).value = "合計"
        ws.cell(row=row, column=4).font = Font(bold=True, name="Yu Gothic UI", size=10)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=5).value = f"=SUM(E2:E{row-1})"
        ws.cell(row=row, column=5).number_format = "#,##0"
        ws.cell(row=row, column=5).font = Font(bold=True, name="Yu Gothic UI", size=10)
        ws.cell(row=row, column=5).fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid")

        for col, w in zip(range(1, 6), [14, 12, 14, 12, 18]):
            set_col_width(ws, col, w)
        ws.freeze_panes = "A2"

        wb.save(os.path.join(branch_dir, f"{branch}.xlsx"))
        print(f"✓ 02_エラーハンドリング演習/支店データ/{branch}.xlsx")

    # ── メインファイル ──
    wb_main = openpyxl.Workbook()
    ws_intro = wb_main.active
    ws_intro.title = "演習説明"
    ws_intro.column_dimensions["A"].width = 3
    ws_intro.column_dimensions["B"].width = 85
    hide_gridlines(ws_intro)

    def put(row, col, val, fnt=None):
        c = ws_intro.cell(row=row, column=col, value=val)
        c.font = fnt or body_font()
        c.alignment = Alignment(vertical="center")
        return c

    put(2,  2, "【演習2】エラーハンドリング", title_font())
    set_row_height(ws_intro, 2, 36)

    put(4,  2, "■ 演習の目的", Font(bold=True, name="Yu Gothic UI", size=11))
    put(5,  2, "支店データフォルダ内のExcelファイルを順番に開き、合計売上を取得するマクロを完成させてください。")
    put(6,  2, "エラーハンドリングを追加して、ファイルが存在しない場合でも止まらないコードにしましょう。")

    put(8,  2, "■ 手順", Font(bold=True, name="Yu Gothic UI", size=11))
    steps = [
        "1. VBAコード_テンプレート.txt を開き、VBEにコードをコピーする",
        "2. 「TODO」コメントが付いた箇所を実装する",
        "3. まずエラーハンドリングなしで実行し、わざとエラーを起こしてみる",
        "   （支店データフォルダから「東京支店.xlsx」を一時的にリネームする）",
        "4. On Error GoTo を追加してエラーハンドリングを実装する",
        "5. ファイルが存在しない場合でも止まらず、「集計結果」シートに結果が出ることを確認する",
    ]
    for i, s in enumerate(steps):
        put(9 + i, 2, s)

    put(16, 2, "■ フォルダ構成", Font(bold=True, name="Yu Gothic UI", size=11))
    tree = [
        "02_エラーハンドリング演習/",
        "    ├ エラーハンドリング演習.xlsx  ← このファイル",
        "    └ 支店データ/",
        "        ├ 東京支店.xlsx",
        "        ├ 大阪支店.xlsx",
        "        └ 名古屋支店.xlsx",
    ]
    for i, t in enumerate(tree):
        c = put(17 + i, 2, t)
        c.font = Font(name="Courier New", size=10, color="555555")

    # 集計結果シート
    ws_res = wb_main.create_sheet("集計結果")
    for col, h in enumerate(["支店名", "合計売上（円）", "取得ステータス"], 1):
        apply_header(ws_res.cell(row=1, column=col), h)
    ws_res.cell(row=2, column=1).value = "（マクロ実行後に表示されます）"
    ws_res.cell(row=2, column=1).font = Font(name="Yu Gothic UI", size=10, color="BBBBBB", italic=True)
    for col, w in zip(range(1, 4), [18, 20, 18]):
        set_col_width(ws_res, col, w)
    set_row_height(ws_res, 1, 25)
    ws_res.freeze_panes = "A2"

    wb_main.save(os.path.join(dir_path, "エラーハンドリング演習.xlsx"))
    print("✓ 02_エラーハンドリング演習/エラーハンドリング演習.xlsx")

    # ── VBAテンプレート ──
    vba = """\
' ============================================================
' 【演習2】エラーハンドリング - 支店データ集計マクロ
' ============================================================
' 「TODO」のコメントが付いた箇所を実装してください。
'
' 【完成イメージ】
' 支店データフォルダ内の Excel ファイルを順番に開き、
' 各支店の合計売上を「集計結果」シートに転記する。
' ファイルが見つからない場合は、エラーメッセージを表示して
' そのファイルをスキップし、次の処理を続ける。
' ============================================================

Sub 支店データ集計()

    Dim folderPath  As String
    Dim fileName    As String
    Dim wb          As Workbook
    Dim ws          As Worksheet
    Dim resultWs    As Worksheet
    Dim resultRow   As Long
    Dim totalSales  As Long
    Dim lastRow     As Long

    ' ----- TODO-1: エラーハンドリングを追加 -----
    ' On Error GoTo ErrorHandler   ← コメントを外してください

    Application.ScreenUpdating = False

    Set resultWs = ThisWorkbook.Worksheets("集計結果")
    resultWs.Range("A2:C100").ClearContents
    resultRow = 2

    folderPath = ThisWorkbook.Path & "\\支店データ\\"

    fileName = Dir(folderPath & "*.xlsx")

    Do While fileName <> ""

        ' TODO-2: ファイルごとにエラーハンドリングをリセットする
        '         ヒント: On Error GoTo ErrorHandler  を書く

        Set wb = Workbooks.Open(folderPath & fileName)
        Set ws = wb.Worksheets("売上データ")

        lastRow    = ws.Cells(ws.Rows.Count, 5).End(xlUp).Row
        totalSales = ws.Cells(lastRow, 5).Value

        resultWs.Cells(resultRow, 1).Value = Replace(fileName, ".xlsx", "")
        resultWs.Cells(resultRow, 2).Value = totalSales
        resultWs.Cells(resultRow, 3).Value = "成功"
        resultRow = resultRow + 1

        wb.Close SaveChanges:=False
        fileName = Dir()

    Loop

    Application.ScreenUpdating = True
    MsgBox "集計が完了しました！" & vbCrLf & _
           "「集計結果」シートを確認してください。"
    Exit Sub

    ' ----- TODO-3: ErrorHandler を実装する -----
    '
    ' ErrorHandler:
    '     ・エラー番号とファイル名をメッセージ表示する
    '     ・結果シートのステータス列に "エラー" と書く
    '     ・ScreenUpdating を True に戻す
    '     ・処理を終了する（または Resume Next でスキップ）

End Sub

' ============================================================
' 参考: On Error の基本パターン
' ============================================================
'
' Sub サンプル()
'     On Error GoTo ErrorHandler
'
'     ' 通常の処理
'     ...
'
'     Exit Sub   ← これがないと ErrorHandler も実行される！
'
' ErrorHandler:
'     MsgBox "エラーが発生しました" & vbCrLf & _
'            "エラー番号: " & Err.Number & vbCrLf & _
'            "内容: "      & Err.Description
'     Application.ScreenUpdating = True
' End Sub
'
' ============================================================
"""
    with open(os.path.join(dir_path, "VBAコード_テンプレート.txt"), "w", encoding="utf-8") as f:
        f.write(vba)
    print("✓ 02_エラーハンドリング演習/VBAコード_テンプレート.txt")


# ===== 演習3: Function演習 =====
def create_function_exercise():
    dir_path = os.path.join(BASE_DIR, "03_Function演習")
    mkdir(dir_path)

    wb = openpyxl.Workbook()

    # ── 演習説明シート ──
    ws_intro = wb.active
    ws_intro.title = "演習説明"
    ws_intro.column_dimensions["A"].width = 3
    ws_intro.column_dimensions["B"].width = 85
    hide_gridlines(ws_intro)

    def put(row, col, val, fnt=None, fill_color=None):
        c = ws_intro.cell(row=row, column=col, value=val)
        c.font = fnt or body_font()
        c.alignment = Alignment(vertical="center")
        if fill_color:
            c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        return c

    put(2, 2, "【演習3】Function（関数）とコードの再利用", title_font())
    set_row_height(ws_intro, 2, 36)

    put(4, 2, "■ 演習の目的", Font(bold=True, name="Yu Gothic UI", size=11))
    put(5, 2, "売上金額をもとに評価（A/B/C）を返す Function を作成し、「売上データ」シートのE列に評価を入力するマクロを完成させてください。")

    put(7, 2, "■ 評価基準", Font(bold=True, name="Yu Gothic UI", size=11))
    criteria = [
        ("A評価", "1,000,000円以上",               OK_COLOR),
        ("B評価", "500,000円以上 ～ 999,999円",     WARN_COLOR),
        ("C評価", "500,000円未満",                  ERR_COLOR),
    ]
    for i, (grade, cond, color) in enumerate(criteria):
        c = put(8 + i, 2, f"  {grade}：{cond}", fill_color=color)
        if i == 0:
            c.font = Font(name="Yu Gothic UI", size=10, bold=True)

    put(12, 2, "■ 手順", Font(bold=True, name="Yu Gothic UI", size=11))
    steps = [
        "1. VBAコード_テンプレート.txt を VBE にコピーする",
        "2. STEP1：評価を返す「Function 売上評価」を完成させる",
        "   （引数: 売上金額 As Long　戻り値: 評価ランク As String）",
        "3. STEP2：最終行を返す「Function 最終行取得」を完成させる",
        "4. STEP3：2つの Function を使って「売上データ」シートのE列に評価を書き込む Sub を完成させる",
        "5. 【確認】E列に A/B/C が正しく入力されていれば完成！",
    ]
    for i, s in enumerate(steps):
        put(13 + i, 2, s)

    # ── 売上データシート ──
    ws_data = wb.create_sheet("売上データ")
    headers = ["担当者名", "地区", "商品カテゴリ", "売上金額（円）", "評価"]
    for col, h in enumerate(headers, 1):
        apply_header(ws_data.cell(row=1, column=col), h)
    set_row_height(ws_data, 1, 25)

    salespeople = [
        ("田中 太郎",  "東京",  "電子機器", 1250000),
        ("鈴木 花子",  "東京",  "家電製品",  820000),
        ("佐藤 一郎",  "大阪",  "電子機器",  450000),
        ("山田 次郎",  "大阪",  "家電製品", 1100000),
        ("中村 美咲",  "名古屋","電子機器",  680000),
        ("小林 健太",  "名古屋","家電製品",  320000),
        ("加藤 三郎",  "福岡",  "電子機器",  950000),
        ("伊藤 由美",  "福岡",  "家電製品", 1380000),
        ("渡辺 誠",    "札幌",  "電子機器",  210000),
        ("松本 志保",  "札幌",  "家電製品",  770000),
        ("木村 拓哉",  "仙台",  "電子機器", 1050000),
        ("林 奈々",    "仙台",  "家電製品",  490000),
    ]

    for row, (name, area, cat, sales) in enumerate(salespeople, 2):
        apply_body(ws_data.cell(row=row, column=1), name,  "left")
        apply_body(ws_data.cell(row=row, column=2), area)
        apply_body(ws_data.cell(row=row, column=3), cat,   "left")
        apply_body(ws_data.cell(row=row, column=4), sales)
        ws_data.cell(row=row, column=4).number_format = "#,##0"

        c = ws_data.cell(row=row, column=5)
        c.value = "（マクロで入力）"
        c.font  = Font(name="Yu Gothic UI", size=10, color="BBBBBB", italic=True)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")

        if row % 2 == 0:
            stripe_row(ws_data, row, 1, 5)
        set_row_height(ws_data, row, 22)

    for col, w in zip(range(1, 6), [16, 10, 14, 18, 10]):
        set_col_width(ws_data, col, w)
    ws_data.freeze_panes = "A2"

    wb.save(os.path.join(dir_path, "売上評価データ.xlsx"))
    print("✓ 03_Function演習/売上評価データ.xlsx")

    # ── VBAテンプレート ──
    vba = """\
' ============================================================
' 【演習3】Function演習 - 売上評価マクロ
' ============================================================
' 3つの STEP に分けて実装してください。
' 各 TODO コメントを参考にしながら進めましょう。
' ============================================================


' ============================================================
' STEP 1: 売上金額から評価ランク（A/B/C）を返す Function を作る
' ============================================================
' 【仕様】
' ・引数:「売上金額」（Long 型）
' ・1,000,000円以上 → "A" を返す
' ・  500,000円以上 → "B" を返す
' ・  500,000円未満 → "C" を返す
' ============================================================

Function 売上評価(売上金額 As Long) As String

    ' TODO: If 文を使って評価を判定し、戻り値として返してください
    '
    ' 【ヒント】Function の戻り値は「関数名 = 値」で返します
    '   例） 売上評価 = "A"

    ' ここに実装してください

End Function


' ============================================================
' STEP 2: ワークシートの最終データ行番号を返す Function を作る
' ============================================================
' 【仕様】
' ・引数:「ワークシート」（Worksheet 型）、「基準列番号」（Long 型）
' ・その列の最終データ行の行番号を返す
' ============================================================

Function 最終行取得(ws As Worksheet, col As Long) As Long

    ' TODO: 最終行番号を返すコードを実装してください
    '
    ' 【ヒント】
    '   ws.Cells(ws.Rows.Count, col).End(xlUp).Row

    ' ここに実装してください

End Function


' ============================================================
' STEP 3: 2つの Function を使い、E列に評価を入力する Sub を作る
' ============================================================

Sub 評価を入力する()

    Dim ws      As Worksheet
    Dim i       As Long
    Dim lastRow As Long
    Dim sales   As Long
    Dim grade   As String

    Set ws = ThisWorkbook.Worksheets("売上データ")

    ' TODO: 「最終行取得」Function を呼び出して lastRow に代入してください
    '       ヒント: 売上金額は D列（列番号 = 4）を基準にします
    lastRow = ' ここに実装してください

    For i = 2 To lastRow

        sales = ws.Cells(i, 4).Value

        ' TODO: 「売上評価」Function を呼び出して grade 変数に代入してください
        grade = ' ここに実装してください

        ws.Cells(i, 5).Value = grade

    Next i

    MsgBox "評価の入力が完了しました！" & vbCrLf & _
           "E列の評価を確認してください。"

End Sub


' ============================================================
' ★ 発展課題（STEP 3 完成後に挑戦してみましょう）
' ============================================================
'
' 1. 評価ごとにセルの背景色を変える
'    A評価 → 緑（RGB: 198, 239, 206）
'    B評価 → 黄（RGB: 255, 235, 156）
'    C評価 → 赤（RGB: 255, 199, 206）
'
' 2. A/B/C それぞれの人数と合計売上をメッセージボックスに表示する
'
' ============================================================
"""
    with open(os.path.join(dir_path, "VBAコード_テンプレート.txt"), "w", encoding="utf-8") as f:
        f.write(vba)
    print("✓ 03_Function演習/VBAコード_テンプレート.txt")


# ===== 演習4: シート操作演習 =====
def create_sheet_exercise():
    dir_path = os.path.join(BASE_DIR, "04_シート操作演習")
    mkdir(dir_path)

    wb = openpyxl.Workbook()

    # ── 演習説明シート ──
    ws_intro = wb.active
    ws_intro.title = "演習説明"
    ws_intro.column_dimensions["A"].width = 3
    ws_intro.column_dimensions["B"].width = 85
    hide_gridlines(ws_intro)

    def put(row, col, val, fnt=None):
        c = ws_intro.cell(row=row, column=col, value=val)
        c.font = fnt or body_font()
        c.alignment = Alignment(vertical="center")
        return c

    put(2, 2, "【演習4】シートの高度な操作", title_font())
    set_row_height(ws_intro, 2, 36)

    put(4, 2, "■ 演習の目的", Font(bold=True, name="Yu Gothic UI", size=11))
    put(5, 2, "「全データ」シートにある1年分の売上データを、月ごとに別シートへ自動で振り分けるマクロを作成してください。")

    put(7, 2, "■ 完成イメージ", Font(bold=True, name="Yu Gothic UI", size=11))
    for i, s in enumerate([
        "・「1月」「2月」…「12月」というシートが自動で作成される",
        "・各月シートには、その月のデータだけが転記されている",
        "・既にシートが存在する場合は削除して再作成する（べき等性）",
    ]):
        put(8 + i, 2, s)

    put(12, 2, "■ 手順", Font(bold=True, name="Yu Gothic UI", size=11))
    for i, s in enumerate([
        "1. VBAコード_テンプレート.txt を VBE にコピーする",
        "2. STEP1：月別シートを作成する Sub を実装する",
        "3. STEP2：「全データ」シートから月ごとにデータを振り分ける Sub を実装する",
        "4. 【確認】各月シートに正しいデータが入っていれば完成！",
    ]):
        put(13 + i, 2, s)

    # ── 全データシート（1年分） ──
    ws_all = wb.create_sheet("全データ")
    headers = ["日付", "支店名", "商品カテゴリ", "担当者", "売上金額（円）"]
    for col, h in enumerate(headers, 1):
        apply_header(ws_all.cell(row=1, column=col), h)
    set_row_height(ws_all, 1, 25)

    branches   = ["東京支店", "大阪支店", "名古屋支店", "福岡支店"]
    categories = ["電子機器", "家電製品", "消耗品", "サービス"]
    staff      = ["田中", "鈴木", "佐藤", "山田", "中村", "小林", "加藤", "伊藤"]
    days_list  = [31,28,31,30,31,30,31,31,30,31,30,31]

    row = 2
    for month in range(1, 13):
        for _ in range(random.randint(8, 14)):
            day    = random.randint(1, days_list[month - 1])
            d      = date(2024, month, day)
            amount = random.randint(5, 200) * 10000

            apply_body(ws_all.cell(row=row, column=1), d.strftime("%Y/%m/%d"))
            apply_body(ws_all.cell(row=row, column=2), random.choice(branches),   "left")
            apply_body(ws_all.cell(row=row, column=3), random.choice(categories), "left")
            apply_body(ws_all.cell(row=row, column=4), random.choice(staff))
            apply_body(ws_all.cell(row=row, column=5), amount)
            ws_all.cell(row=row, column=5).number_format = "#,##0"
            if row % 2 == 0:
                stripe_row(ws_all, row, 1, 5)
            set_row_height(ws_all, row, 22)
            row += 1

    for col, w in zip(range(1, 6), [14, 14, 14, 10, 18]):
        set_col_width(ws_all, col, w)
    ws_all.freeze_panes = "A2"

    wb.save(os.path.join(dir_path, "月次売上データ.xlsx"))
    print("✓ 04_シート操作演習/月次売上データ.xlsx")

    # ── VBAテンプレート ──
    vba = """\
' ============================================================
' 【演習4】シート操作演習 - 月別シート自動振り分けマクロ
' ============================================================
' 「全データ」シートの売上データを月ごとに振り分けて
' 月別シートに転記するマクロを完成させてください。
'
' 【データ構造】全データシート
'   A列: 日付（例: 2024/01/15）
'   B列: 支店名
'   C列: 商品カテゴリ
'   D列: 担当者
'   E列: 売上金額
' ============================================================


' ============================================================
' STEP 1: 月別シートを準備する Sub
' ============================================================
' 1月～12月のシートを作成する。
' 既に同名シートが存在する場合は削除してから作成する。

Sub 月別シートを準備する()

    Dim i         As Long
    Dim sheetName As String
    Dim ws        As Worksheet

    Application.DisplayAlerts = False   ' 削除時の確認ダイアログを非表示

    For i = 1 To 12
        sheetName = i & "月"

        ' TODO: シートが既に存在する場合は削除してください
        '
        ' 【ヒント】以下のパターンでシートの存在確認ができます
        '   On Error Resume Next
        '   Set ws = Worksheets(sheetName)
        '   On Error GoTo 0
        '   If Not ws Is Nothing Then
        '       ws.Delete
        '       Set ws = Nothing
        '   End If

        ' ここに実装してください


        ' TODO: 新しいシートを末尾に追加し、シート名を設定してください
        '
        ' 【ヒント】
        '   Worksheets.Add(After:=Worksheets(Worksheets.Count)).Name = sheetName

        ' ここに実装してください


        ' 追加したシートにヘッダーを設定する
        With Worksheets(sheetName)
            .Cells(1, 1).Value = "日付"
            .Cells(1, 2).Value = "支店名"
            .Cells(1, 3).Value = "商品カテゴリ"
            .Cells(1, 4).Value = "担当者"
            .Cells(1, 5).Value = "売上金額（円）"
        End With

    Next i

    Application.DisplayAlerts = True
    MsgBox "月別シートの準備が完了しました！（1月～12月）"

End Sub


' ============================================================
' STEP 2: データを月別シートに振り分ける Sub
' ============================================================

Sub データを月別に振り分ける()

    Dim srcWs       As Worksheet  ' 全データシート
    Dim dstWs       As Worksheet  ' 転記先シート
    Dim i           As Long
    Dim lastRow     As Long
    Dim dstLastRow  As Long
    Dim dateVal     As Date
    Dim monthNum    As Long
    Dim sheetName   As String

    Application.ScreenUpdating = False

    Set srcWs = ThisWorkbook.Worksheets("全データ")
    lastRow = srcWs.Cells(srcWs.Rows.Count, 1).End(xlUp).Row

    For i = 2 To lastRow

        dateVal   = srcWs.Cells(i, 1).Value
        monthNum  = Month(dateVal)          ' Month 関数で月番号を取得
        sheetName = monthNum & "月"

        ' TODO: 転記先シートを取得し、次の空行に A～E 列を転記してください
        '
        ' 【ヒント】
        '   Set dstWs = Worksheets(sheetName)
        '   dstLastRow = dstWs.Cells(dstWs.Rows.Count, 1).End(xlUp).Row + 1
        '   で転記先の次の空行が分かります
        '
        '   srcWs の i 行目の値を dstWs の dstLastRow 行目にコピーしてください
        '   （A列～E列の5列分）

        ' ここに実装してください


    Next i

    Application.ScreenUpdating = True
    MsgBox "振り分けが完了しました！" & vbCrLf & _
           "各月シートのデータ件数を確認してください。"

End Sub


' ============================================================
' まとめて実行する Sub
' ============================================================

Sub 月別振り分け実行()
    Call 月別シートを準備する
    Call データを月別に振り分ける
End Sub


' ============================================================
' ★ 発展課題
' ============================================================
'
' 1. 各月シートのヘッダー行に背景色をつける
' 2. 各月シートの最終行に合計売上を表示する
' 3. E列に数値書式（#,##0）を設定する
'
' ============================================================
"""
    with open(os.path.join(dir_path, "VBAコード_テンプレート.txt"), "w", encoding="utf-8") as f:
        f.write(vba)
    print("✓ 04_シート操作演習/VBAコード_テンプレート.txt")


# ===== 演習5: 総合演習 =====
def create_comprehensive_exercise():
    dir_path   = os.path.join(BASE_DIR, "05_総合演習")
    branch_dir = os.path.join(dir_path, "支店データ")
    mkdir(dir_path)
    mkdir(branch_dir)

    branches = ["東京支店", "大阪支店", "名古屋支店", "福岡支店"]
    months   = list(range(1, 7))   # 1月〜6月
    products = {
        "電子機器A": 45000, "電子機器B": 28000,
        "家電製品A": 15000, "家電製品B":  8500,
        "消耗品A":    3200, "消耗品B":    1800,
    }
    staff_map = {
        "東京支店":   ["田中 太郎", "鈴木 花子", "佐藤 一郎", "山田 次郎"],
        "大阪支店":   ["中村 美咲", "小林 健太", "加藤 三郎", "伊藤 由美"],
        "名古屋支店": ["渡辺 誠",   "松本 志保", "木村 拓哉", "林 奈々"],
        "福岡支店":   ["高橋 隆",   "岡田 恵子", "前田 秀樹", "藤田 彩"],
    }
    days_list = [31,28,31,30,31,30]

    for branch in branches:
        for month in months:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "売上明細"

            headers = ["日付", "商品名", "担当者", "数量（個）", "単価（円）", "売上金額（円）"]
            for col, h in enumerate(headers, 1):
                apply_header(ws.cell(row=1, column=col), h)
            set_row_height(ws, 1, 25)

            staff        = staff_map[branch]
            product_names = list(products.keys())
            row = 2

            for _ in range(random.randint(10, 18)):
                day   = random.randint(1, days_list[month - 1])
                d     = date(2024, month, day)
                prod  = random.choice(product_names)
                price = products[prod]
                qty   = random.randint(1, 30)
                amt   = qty * price

                apply_body(ws.cell(row=row, column=1), d.strftime("%Y/%m/%d"))
                apply_body(ws.cell(row=row, column=2), prod, "left")
                apply_body(ws.cell(row=row, column=3), random.choice(staff), "left")
                apply_body(ws.cell(row=row, column=4), qty)
                apply_body(ws.cell(row=row, column=5), price)
                apply_body(ws.cell(row=row, column=6), amt)
                ws.cell(row=row, column=5).number_format = "#,##0"
                ws.cell(row=row, column=6).number_format = "#,##0"
                if row % 2 == 0:
                    stripe_row(ws, row, 1, 6)
                set_row_height(ws, row, 22)
                row += 1

            # 合計行
            ws.cell(row=row, column=5).value = "月間合計"
            ws.cell(row=row, column=5).font  = Font(bold=True, name="Yu Gothic UI", size=10)
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=6).value  = f"=SUM(F2:F{row-1})"
            ws.cell(row=row, column=6).number_format = "#,##0"
            ws.cell(row=row, column=6).font   = Font(bold=True, name="Yu Gothic UI", size=10)
            ws.cell(row=row, column=6).fill   = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid")

            for col, w in zip(range(1, 7), [14, 16, 14, 12, 14, 18]):
                set_col_width(ws, col, w)
            ws.freeze_panes = "A2"

            filename = f"{branch}_{2024}{month:02d}.xlsx"
            wb.save(os.path.join(branch_dir, filename))
            print(f"✓ 05_総合演習/支店データ/{filename}")

    # ── ベースファイル ──
    wb_main  = openpyxl.Workbook()
    ws_intro = wb_main.active
    ws_intro.title = "演習説明"
    ws_intro.column_dimensions["A"].width = 3
    ws_intro.column_dimensions["B"].width = 90
    hide_gridlines(ws_intro)

    def put(row, col, val, fnt=None):
        c = ws_intro.cell(row=row, column=col, value=val)
        c.font = fnt or body_font()
        c.alignment = Alignment(vertical="center")
        return c

    put(2, 2, "【演習5】総合演習 ─ 壊れにくい実務ツールを作る", title_font())
    set_row_height(ws_intro, 2, 36)

    put(4, 2, "■ 演習の目的", Font(bold=True, name="Yu Gothic UI", size=11))
    put(5, 2, "1日目の「複数ファイル統合マクロ」を発展させ、2日目で学んだすべての要素を組み込んだ実務ツールを完成させます。")

    put(7, 2, "■ 完成イメージ", Font(bold=True, name="Yu Gothic UI", size=11))
    for i, s in enumerate([
        "1. ボタンを押す",
        "2. 処理開始前に画面更新・自動計算を停止",
        "3. 支店データフォルダ内のファイルを順番に開く",
        "4.   ★ ファイルが開けなかった → エラーメッセージを表示してスキップ（エラーハンドリング）",
        "5.   最終行を Function で取得し、月間合計を読み取る（Function の活用）",
        "6.   集計結果シートに転記する",
        "7.   ファイルを閉じる",
        "8. 処理完了メッセージを表示",
        "9. 画面更新・自動計算を元に戻す（後片付け）",
    ]):
        put(8 + i, 2, s)

    put(18, 2, "■ ステップ別進行", Font(bold=True, name="Yu Gothic UI", size=11))
    for i, s in enumerate([
        "Step 1【必須】: エラーハンドリングを追加する",
        "Step 2【必須】: 最終行取得を Function に切り出す",
        "Step 3【推奨】: ScreenUpdating 等で処理を最適化する",
        "Step 4【発展】: 集計結果シートに書式を設定する",
        "Step 5【発展】: デバッグしながら動作確認・仕上げ",
    ]):
        put(19 + i, 2, s)

    put(25, 2, "■ フォルダ構成", Font(bold=True, name="Yu Gothic UI", size=11))
    tree = [
        "05_総合演習/",
        "    ├ 集計ツール_ベース.xlsx  ← このファイル",
        "    └ 支店データ/",
        "        ├ 東京支店_202401.xlsx",
        "        ├ 東京支店_202402.xlsx  … （1月〜6月）",
        "        ├ 大阪支店_202401.xlsx",
        "        ├ 名古屋支店_202401.xlsx",
        "        └ 福岡支店_202406.xlsx  ← 計24ファイル（4支店 × 6ヶ月）",
    ]
    for i, t in enumerate(tree):
        c = put(26 + i, 2, t)
        c.font = Font(name="Courier New", size=10, color="555555")

    # 集計結果シート
    ws_res = wb_main.create_sheet("集計結果")
    for col, h in enumerate(["支店名", "年月", "月間売上合計（円）", "明細件数", "ステータス"], 1):
        apply_header(ws_res.cell(row=1, column=col), h)
    ws_res.cell(row=2, column=1).value = "（マクロ実行後に表示されます）"
    ws_res.cell(row=2, column=1).font  = Font(name="Yu Gothic UI", size=10, color="BBBBBB", italic=True)
    for col, w in zip(range(1, 6), [16, 12, 22, 12, 18]):
        set_col_width(ws_res, col, w)
    set_row_height(ws_res, 1, 25)
    ws_res.freeze_panes = "A2"

    wb_main.save(os.path.join(dir_path, "集計ツール_ベース.xlsx"))
    print("✓ 05_総合演習/集計ツール_ベース.xlsx")

    # ── VBAテンプレート ──
    vba = """\
' ============================================================
' 【演習5】総合演習 - 支店別売上集計ツール
' ============================================================
' 2日目で学んだすべての要素を組み合わせて
' 実務で使えるツールを完成させましょう。
'
' ファイル名形式: 支店名_YYYYMM.xlsx（例: 東京支店_202401.xlsx）
' 各ファイルの「売上明細」シートの最終行 F 列に月間合計があります
'
' 【TODO 一覧】
' [必須] TODO-1: エラーハンドリングの追加
' [必須] TODO-2: Function の呼び出し（最終行取得）
' [推奨] TODO-3: ScreenUpdating 等の最適化
' [発展] TODO-4: 集計結果シートへの書式設定
' ============================================================


' ============================================================
' 共通 Function
' ============================================================

' 最終行番号を返す Function
Function 最終行取得(ws As Worksheet, col As Long) As Long
    最終行取得 = ws.Cells(ws.Rows.Count, col).End(xlUp).Row
End Function

' ファイル名から支店名を返す Function
' 例: "東京支店_202401.xlsx" → "東京支店"
Function 支店名を取得(fileName As String) As String
    Dim parts() As String
    parts = Split(Replace(fileName, ".xlsx", ""), "_")
    支店名を取得 = parts(0)
End Function

' ファイル名から年月文字列を返す Function
' 例: "東京支店_202401.xlsx" → "2024年01月"
Function 年月を取得(fileName As String) As String
    Dim parts() As String
    parts = Split(Replace(fileName, ".xlsx", ""), "_")
    年月を取得 = Left(parts(1), 4) & "年" & Right(parts(1), 2) & "月"
End Function


' ============================================================
' メインプロシージャ
' ============================================================

Sub 支店別売上集計()

    Dim folderPath  As String
    Dim fileName    As String
    Dim wb          As Workbook
    Dim srcWs       As Worksheet
    Dim resultWs    As Worksheet
    Dim resultRow   As Long
    Dim lastRow     As Long
    Dim totalSales  As Long
    Dim recordCount As Long

    ' ----- TODO-3: 最適化設定（コメントを外してください）-----
    ' Application.ScreenUpdating = False
    ' Application.Calculation = xlCalculationManual
    ' Application.DisplayAlerts = False

    ' ----- TODO-1: エラーハンドリングを追加（コメントを外してください）-----
    ' On Error GoTo ErrorHandler

    ' ----- 初期化 -----
    Set resultWs = ThisWorkbook.Worksheets("集計結果")
    resultWs.Range("A2:E1000").ClearContents
    resultRow = 2

    folderPath = ThisWorkbook.Path & "\\支店データ\\"

    ' ----- ファイル処理ループ -----
    fileName = Dir(folderPath & "*.xlsx")

    Do While fileName <> ""

        On Error GoTo FileError   ' ファイルごとのエラーを個別にキャッチ

        Set wb    = Workbooks.Open(folderPath & fileName, ReadOnly:=True)
        Set srcWs = wb.Worksheets("売上明細")

        ' TODO-2: 最終行取得 Function を使って lastRow に代入してください
        '         ヒント: lastRow = 最終行取得(srcWs, 1)
        lastRow = ' ここに実装してください

        ' 月間合計（最終行の F 列）を取得
        totalSales  = srcWs.Cells(lastRow, 6).Value
        recordCount = lastRow - 1   ' ヘッダー除く明細件数

        ' 集計結果シートに転記
        resultWs.Cells(resultRow, 1).Value = 支店名を取得(fileName)
        resultWs.Cells(resultRow, 2).Value = 年月を取得(fileName)
        resultWs.Cells(resultRow, 3).Value = totalSales
        resultWs.Cells(resultRow, 4).Value = recordCount
        resultWs.Cells(resultRow, 5).Value = "成功"
        resultRow = resultRow + 1

        wb.Close SaveChanges:=False
        Set wb = Nothing

        On Error GoTo ErrorHandler  ' 個別ハンドラーを解除

        fileName = Dir()

    Loop

    ' ----- 正常終了 -----
    Call 後片付け
    MsgBox "集計が完了しました！" & vbCrLf & _
           (resultRow - 2) & " 件のファイルを処理しました。", vbInformation, "完了"
    Exit Sub


' ----- ファイル個別エラー（スキップして続行）-----
FileError:
    If Not wb Is Nothing Then
        wb.Close SaveChanges:=False
        Set wb = Nothing
    End If
    resultWs.Cells(resultRow, 1).Value = 支店名を取得(fileName)
    resultWs.Cells(resultRow, 2).Value = 年月を取得(fileName)
    resultWs.Cells(resultRow, 5).Value = "エラー: " & Err.Description
    resultRow = resultRow + 1
    Resume Next   ' 次のループ処理へ


' ----- 全体エラー（処理を中止）-----
ErrorHandler:
    MsgBox "予期しないエラーが発生しました。" & vbCrLf & _
           "エラー番号: " & Err.Number & vbCrLf & _
           "内容: "       & Err.Description, vbCritical, "エラー"
    Call 後片付け

End Sub


' ============================================================
' 後片付け処理（正常終了・エラー終了の両方から呼ぶ）
' ============================================================

Sub 後片付け()
    Application.ScreenUpdating = True
    Application.Calculation    = xlCalculationAutomatic
    Application.DisplayAlerts  = True
End Sub


' ============================================================
' ★ 発展課題
' ============================================================
'
' 1. 集計結果シートに書式を設定する
'    ・ヘッダー行に背景色をつける
'    ・売上金額列に数値書式（#,##0）を設定する
'    ・成功 → 緑、エラー → 赤 で文字色を変える
'
' 2. 処理時間をメッセージに表示する
'    （Timer 関数で開始・終了時刻を記録して差分を計算）
'
' 3. 支店ごとの合計をサマリーシートに別途まとめる
'
' ============================================================
"""
    with open(os.path.join(dir_path, "VBAコード_テンプレート.txt"), "w", encoding="utf-8") as f:
        f.write(vba)
    print("✓ 05_総合演習/VBAコード_テンプレート.txt")


# ===== メイン =====
def main():
    print("=" * 55)
    print("  2日目研修 演習用ファイル 生成開始")
    print("=" * 55)

    mkdir(BASE_DIR)

    create_debug_exercise()
    print()
    create_error_handling_exercise()
    print()
    create_function_exercise()
    print()
    create_sheet_exercise()
    print()
    create_comprehensive_exercise()

    print()
    print("=" * 55)
    print("  ✅ 全ファイルの生成が完了しました")
    print(f"  保存先: {BASE_DIR}")
    print("=" * 55)


if __name__ == "__main__":
    main()
